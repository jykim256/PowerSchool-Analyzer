import os.path,os,subprocess,sys,string,time,openpyxl
from cleanTableR import loadData
from openpyxl.styles import Alignment, Font
from openpyxl.chart import BarChart, Reference,Series

# include the grades that teachers exclude?
exclude = True
notRecord = ['Excluded','Exempt']
catIndex = 1
morgoVariable = False
sortCategory = False

def assignCell(cell,value,font,alignment):
	cell.value = value
	cell.font = font
	cell.alignment = alignment

def classAnalysis(grades,className):
	period = grades
	excludeIndex = 7
	numLetter = 'D'
	denLetter = 'E'

	# count number of grading categories there are
	allCat = [period[i][catIndex] for i in range(len(period))]
	trueCat = list(map(lambda x: allCat[x] in allCat[0:x], range(len(period))))
	numCat = len(trueCat) - sum(trueCat)
	categories = ['' for _ in range(numCat)]
	k = 0
	for i in range(len(allCat)):
		if not(trueCat[i]):
			categories[k] = allCat[i]
			k += 1

	# separates assignments by category and records the indices
	catIndices = [[] for _ in range(numCat)]
	for i in range(len(period)):
		for j in range(len(categories)):
			if(period[i][catIndex] == categories[j]):
				# appends the indexes
				if exclude:
					try:
						float(period[i][5])
						if not(period[i][excludeIndex] in notRecord):
							catIndices[j].append(i)
					except:
						pass
				else:
					catIndices[j].append(i)

	num = ['' for _ in range(numCat)]; den = ['' for _ in range(numCat)]
	for i in range(len(catIndices)):
		a = '='
		b = '='
		for j in catIndices[i]:
			# add one bc Excel is 1 index system, then another bc top row adds one
			a += denLetter + str(j+2) + '+'
			b += numLetter + str(j+2) + '+'
		den[i] = a[:-1]; num[i] = b[:-1]
	return [categories,num,den]

def exportXL(currentGrades,grades,classes,XLName):
	for i in range(len(grades)):
		grades[i] = grades[i].tolist()
	# tries to convert everything to floats
	for i in range(len(grades)):
		for j in range(len(grades[i])):
			for k in range(len(grades[i][j])):
				try:
					grades[i][j][k] = float(grades[i][j][k])
				except:
					pass

	# finds the number of actual classes
	numClasses = 0
	for i in range(len(classes)):
		try:
			if classes[i][0]:
				numClasses += 1
		except:
			pass

	# sorts by the category
	if sortCategory:
		for i in grades:
			try:
				i.sort(key=lambda x: x[catIndex])
			except:
				pass

	aIndex = ord('A')
	book = ''
	try:
		book = openpyxl.load_workbook(filename = XLName)
		for i in range(numClasses):
			sheet = book[classes[i][0]]
	except:
		print('Book not found')
		book = openpyxl.Workbook()
		for i in range(numClasses):
			# makes a new sheet for each class
			if i == 0:
				sheet = book.active
				sheet.title = classes[0][0]
			else:
				sheet = book.create_sheet(classes[i][0])
		print('New book generated...')

	ftBold = Font(name = 'Roboto',bold = True)
	ftReg = Font(name = 'Roboto',bold = False)
	alCenter = Alignment(horizontal = 'center',vertical = 'center')
	alRight = Alignment(horizontal = 'right',vertical = 'center')
	alLeft = Alignment(horizontal = 'left',vertical = 'center')

	# for each class a different sheet
	for i in range(numClasses):
		sheet = book[classes[i][0]]

		# creates column titles for assignments
		colTitles = ['Date','Category','Assignment Name','Earned','Possible','Percentage','Grade','Comments']
		for j in range(len(colTitles)):
			cell = sheet[chr(aIndex + j) + str(1)]
			cell.value = colTitles[j]
			cell.font = ftBold
			cell.alignment = alCenter

		# makes a new row for each assignment
		for assNum in range(len(grades[i])):
			for col in range(len(grades[i][assNum])):
				cell = sheet[chr(aIndex + col) + str(2 + assNum)]
				cell.value = grades[i][assNum][col]
				cell.font = ftReg
				if col == catIndex + 1:	
					cell.alignment = alLeft
				else:
					cell.alignment = alCenter

		# gets the category scores and gets them out
		categories,num,den = classAnalysis(grades[i],classes[i][0])

		rowTitles = ['Category Name','Points Earned','Possible Points', 'Category Percentage','Category Weight','1 Pt is worth']
		maxNumCols = len(grades[i][assNum])

		tableOffset = 2

		# prints the row titles for category printing
		for row in range(len(rowTitles)):
			cell = sheet[chr(aIndex + maxNumCols + tableOffset - 1) + str(row + 1)]
			cell.value = rowTitles[row]
			cell.font = ftBold
			cell.alignment = alRight

		s = '=round('
		for col in range(len(categories)):
			colIndex = aIndex + col + maxNumCols + tableOffset
			# print the category names
			cell = sheet[chr(colIndex) + str(1)]
			assignCell(cell,categories[col],ftBold,alCenter)
			# print points earned
			cell = sheet[chr(colIndex) + str(2)]
			assignCell(cell,num[col],ftReg,alCenter)
			# print points possible
			cell = sheet[chr(colIndex) + str(3)]
			assignCell(cell,den[col],ftReg,alCenter)
			# print category percentage
			cell = sheet[chr(colIndex) + str(4)]
			assignCell(cell,'=round(' + chr(colIndex) + '2/' + chr(colIndex) + '3,3) * 100',ftReg,alCenter)
			# print percentage per point
			cell = sheet[chr(colIndex) + str(6)]
			assignCell(cell,'=round( 1 / (' + chr(colIndex) + '3+5)*' + chr(colIndex) + '5,3)',ftReg,alCenter)
			# adds the category percentage * category weight
			s += chr(colIndex) + '4 * ' + chr(colIndex) + '5 + '
		# prints label for grade
		cell = sheet[chr(aIndex + len(categories) + maxNumCols + tableOffset - 2) + str(8)]
		assignCell(cell,'Grades:',ftBold,alCenter)
		# prints grade
		cell = sheet[chr(aIndex + len(categories) + maxNumCols + tableOffset - 1) + str(8)]
		assignCell(cell,s[:-3] + ',3) / 100',ftBold,alCenter)

		cell = sheet[chr(aIndex + len(categories) + maxNumCols + tableOffset - 2) + str(9)]
		assignCell(cell,'WebGrades:',ftBold,alCenter)
		# prints grade
		cell = sheet[chr(aIndex + len(categories) + maxNumCols + tableOffset - 1) + str(9)]
		assignCell(cell,str(currentGrades[i]) + '%',ftBold,alCenter)


		chart = BarChart()
		chart.type = "col"
		chart.style = 2
		chart.title = "Category Percentage"
		chart.x_axis.title = 'Categories'
		chart.y_axis.title = 'Percent (%)'
		chart.x_axis.delete = True
		chart.y_axis.scaling.min = 50
		chart.y_axis.scaling.max = 101

		start = maxNumCols + tableOffset + 1
		end = start+len(categories) - 1
		cats = Reference(sheet, min_col=start, max_col=end, min_row=1, max_row=1)
		for j in range(start,end+1):
			data = Reference(sheet, min_col=j, min_row=4, max_row=4)
			series = Series(data, title = categories[j - start])
			chart.series.append(series)

		chart.grouping = 'clustered'
		sheet.add_chart(chart, 'K15')

	isOpen = True
	while(isOpen):
		try:
			book.save(XLName)
			isOpen = False
		except:
			#os.system('taskkill /F /IM excel.exe')
			print('Close Excel Spreadsheet!')
			time.sleep(0.8)
	print(XLName + ' updated.')
	input('Press [Enter] to open Excel File\n')
	os.system(XLName)
	#sys.exit()

def getXLName():
	return 'Grades.xlsx'


if __name__ == "__main__":

	fileName = 'PSData.npz'
	XLName = getXLName()

	#checks if file exists
	if os.path.isfile(fileName):
		print('Previous data exists!')
	# gets the variables
	a = loadData(fileName,[''])
	grades = a[1]
	classes = a[2]
	print(classes)
	currentGrades = [454,540,30,20]
	exportXL(currentGrades,grades,classes,XLName)